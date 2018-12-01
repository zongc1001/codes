import openpyxl
filename = 'UserBehavior'
inwb = openpyxl.load_workbook(filename)
sheetname = inwb.get_sheet_names()

ws = inwb.get_sheet_by_name(sheetname[0])
rows = ws.max_row
cols = ws.max_column
print(ws.cell(1, 1).value)